import os
import pandas as pd
import numpy as np
import warnings
from pathlib import Path
warnings.filterwarnings("ignore")

# ==================================================
# CONFIGURATION
# ==================================================
DATA_FOLDER = "data_folder"  # Generic folder name - place CSV files here
OUTPUT_FILE = "compiled_analysis_report.xlsx"

# ==================================================
# HELPER FUNCTIONS
# ==================================================
def extract_base_instrument(filename):
    """
    Extracts the base instrument name from the CSV filename.
    Example: 'FUTIDX_BANKNIFTY_01-Apr-2024_TO_30-Jun-2024.csv' -> 'BANKNIFTY'
    Example: 'FUTSTK_TATAMOTORS_01-Jan-2022_TO_31-Mar-2022.csv' -> 'TATAMOTORS'
    """
    # Remove .csv extension
    name = filename.replace('.csv', '')
    
    # Split by underscore
    parts = name.split('_')
    
    # The instrument name is the second part (after FUTIDX or FUTSTK)
    if len(parts) >= 2:
        return parts[1]
    
    return name

# ==================================================
# ANALYZE ONE CSV (CORE LOGIC)
# ==================================================
def analyze_single_dataframe(df_raw, instrument_name, base_instrument, oi_floor=2000):
    """
    Analyzes a single futures CSV file.
    Returns cleaned, contract-separated dataframe with all signals.
    """
    df = df_raw.copy()
    df.columns = df.columns.str.strip()

    # Validate required columns
    if "Date" not in df.columns:
        return pd.DataFrame()

    # Parse dates
    df["Date"] = pd.to_datetime(df["Date"], format="%d-%b-%Y", errors="coerce")

    # Contract identification
    if "Expiry" in df.columns:
        df["Contract_ID"] = df["Expiry"]
    elif "Expiry_Date" in df.columns:
        df["Contract_ID"] = df["Expiry_Date"]
    else:
        df["Contract_ID"] = df.index.astype(str)

    # Numeric conversion
    numeric_cols = ["Open", "Close", "No. of contracts"]
    if "Open Int" in df.columns:
        numeric_cols.append("Open Int")

    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Drop invalid rows
    required_cols = ["Open", "Close", "No. of contracts", "Date"]
    df = df.dropna(subset=required_cols)
    df = df.sort_values(["Date", "Contract_ID"])

    # Process each contract separately
    all_contracts = []

    for cid in df["Contract_ID"].unique():
        cdf = df[df["Contract_ID"] == cid].copy()
        cdf = cdf.sort_values("Date").reset_index(drop=True)

        # Skip contracts with insufficient data
        if len(cdf) <= 6:
            continue

        # Remove rollover noise (first 3 and last 3 days)
        cdf = cdf.iloc[3:-3].reset_index(drop=True)

        # OI floor filter (if OI exists)
        if "Open Int" in cdf.columns:
            cdf = cdf[cdf["Open Int"].shift(1) >= oi_floor]

        if len(cdf) < 2:
            continue

        # === CORE SIGNALS ===
        # Price signals
        cdf["Daily_Change"] = cdf["Close"] - cdf["Open"]
        cdf["Is_Loss"] = cdf["Daily_Change"] < 0
        cdf["Is_Gain"] = cdf["Daily_Change"] > 0

        # Volume signals
        cdf["Volume_Pct_Change"] = cdf["No. of contracts"].pct_change() * 100

        # OI signals (if available)
        if "Open Int" in cdf.columns:
            cdf["OI_Change"] = cdf["Open Int"].diff()
            cdf["OI_20D_Avg"] = cdf["Open Int"].rolling(20, min_periods=5).mean()
            cdf["OI_Normalized_Change"] = cdf["OI_Change"] / cdf["OI_20D_Avg"]

            # Next-day OI response
            cdf["Next_Day_OI"] = cdf["Open Int"].shift(-1)
            cdf["Next_Day_OI_Change"] = cdf["Next_Day_OI"] - cdf["Open Int"]
            cdf["Next_Day_OI_Normalized_Change"] = (
                cdf["Next_Day_OI_Change"] / cdf["OI_20D_Avg"]
            )

        # Next-day volume response
        cdf["Next_Day_Volume"] = cdf["No. of contracts"].shift(-1)
        cdf["Next_Day_Volume_Pct_Change"] = (
            (cdf["Next_Day_Volume"] - cdf["No. of contracts"]) 
            / cdf["No. of contracts"] * 100
        )

        # Add instrument identifiers (both full name and base name)
        cdf["Instrument"] = instrument_name
        cdf["Base_Instrument"] = base_instrument
        cdf["Contract_ID"] = cid

        all_contracts.append(cdf)

    if not all_contracts:
        return pd.DataFrame()

    return pd.concat(all_contracts, ignore_index=True)


# ==================================================
# ANALYZE ENTIRE FOLDER
# ==================================================
def analyze_data_folder(data_folder_path=DATA_FOLDER, oi_floor=2000):
    """
    Analyzes ALL CSV files in the data folder.
    Returns aggregated dataframe with all instruments and contracts.
    """
    if not os.path.exists(data_folder_path):
        raise FileNotFoundError(f"Data folder not found: {data_folder_path}")

    all_data = []
    files_processed = 0
    files_failed = 0

    print(f"Starting analysis of folder: {data_folder_path}")
    print("=" * 60)

    # Get all CSV files in the folder
    csv_files = [f for f in os.listdir(data_folder_path) if f.lower().endswith(".csv")]
    
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in {data_folder_path}")
    
    print(f"Found {len(csv_files)} CSV file(s) to process\n")

    for filename in csv_files:
        file_path = os.path.join(data_folder_path, filename)
        instrument_name = filename.replace(".csv", "")
        base_instrument = extract_base_instrument(filename)

        try:
            df_raw = pd.read_csv(file_path)
            df_analyzed = analyze_single_dataframe(df_raw, instrument_name, base_instrument, oi_floor)

            if not df_analyzed.empty:
                all_data.append(df_analyzed)
                files_processed += 1
                print(f"✓ Processed: {filename} ({len(df_analyzed)} rows) -> Base: {base_instrument}")
            else:
                files_failed += 1
                print(f"✗ Skipped: {filename} (insufficient data)")

        except Exception as e:
            files_failed += 1
            print(f"✗ Failed: {filename} - {str(e)}")

    print("=" * 60)
    print(f"Summary: {files_processed} files processed, {files_failed} files skipped/failed")

    if not all_data:
        raise ValueError("No valid data found in any CSV files")

    # Combine all instruments and contracts
    df_combined = pd.concat(all_data, ignore_index=True)
    return df_combined



# ==================================================
# GENERATE COMPILED REPORT
# ==================================================
def generate_compiled_report(df_combined):
    """
    Generates contract-level compiled report with instrument-wise and overall averages.
    Now includes percentile rankings for OI changes.
    """
    df_work = df_combined.copy()
    
    # Calculate percentiles for OI changes (global across all data)
    if "Next_Day_OI_Normalized_Change" in df_work.columns:
        df_work["Next_Day_OI_Pctl"] = (
            df_work["Next_Day_OI_Normalized_Change"].rank(pct=True) * 100
        )
    
    report_data = []
    
    # Group by instrument and contract
    for (instrument, contract_id), group in df_work.groupby(['Instrument', 'Contract_ID']):
        loss_days = group[group['Is_Loss']]
        gain_days = group[group['Is_Gain']]
        
        row = {
            'Instrument': instrument,
            'Base_Instrument': group['Base_Instrument'].iloc[0],
            'Contract_ID': contract_id,
            'Total_Days': len(group),
            'Loss_Days': len(loss_days),
            'Gain_Days': len(gain_days),
        }
        
        # After-loss metrics with percentiles
        if len(loss_days) > 0:
            row['Avg_OI_Percentile_AfterLoss'] = loss_days['Next_Day_OI_Pctl'].mean() if 'Next_Day_OI_Pctl' in loss_days.columns else np.nan
            row['Avg_NextDay_Volume_Change_AfterLoss'] = loss_days['Next_Day_Volume_Pct_Change'].mean()
            row['Avg_NextDay_OI_Normalized_AfterLoss'] = loss_days['Next_Day_OI_Normalized_Change'].mean() if 'Next_Day_OI_Normalized_Change' in loss_days.columns else np.nan
            row['Pct_OI_Increase_AfterLoss'] = (loss_days['Next_Day_OI_Change'] > 0).mean() * 100 if 'Next_Day_OI_Change' in loss_days.columns else np.nan
        else:
            row['Avg_OI_Percentile_AfterLoss'] = np.nan
            row['Avg_NextDay_Volume_Change_AfterLoss'] = np.nan
            row['Avg_NextDay_OI_Normalized_AfterLoss'] = np.nan
            row['Pct_OI_Increase_AfterLoss'] = np.nan
        
        # After-gain metrics with percentiles
        if len(gain_days) > 0:
            row['Avg_OI_Percentile_AfterGain'] = gain_days['Next_Day_OI_Pctl'].mean() if 'Next_Day_OI_Pctl' in gain_days.columns else np.nan
            row['Avg_NextDay_Volume_Change_AfterGain'] = gain_days['Next_Day_Volume_Pct_Change'].mean()
            row['Avg_NextDay_OI_Normalized_AfterGain'] = gain_days['Next_Day_OI_Normalized_Change'].mean() if 'Next_Day_OI_Normalized_Change' in gain_days.columns else np.nan
            row['Pct_OI_Increase_AfterGain'] = (gain_days['Next_Day_OI_Change'] > 0).mean() * 100 if 'Next_Day_OI_Change' in gain_days.columns else np.nan
        else:
            row['Avg_OI_Percentile_AfterGain'] = np.nan
            row['Avg_NextDay_Volume_Change_AfterGain'] = np.nan
            row['Avg_NextDay_OI_Normalized_AfterGain'] = np.nan
            row['Pct_OI_Increase_AfterGain'] = np.nan
        
        report_data.append(row)
    
    df_report = pd.DataFrame(report_data)
    df_report = df_report.sort_values(['Base_Instrument', 'Contract_ID'])
    
    # Calculate instrument-wise averages
    instrument_averages = []
    for base_instrument in df_report['Base_Instrument'].unique():
        inst_data = df_report[df_report['Base_Instrument'] == base_instrument]
        
        avg_row = {
            'Instrument': f"{base_instrument} - AVERAGE",
            'Base_Instrument': base_instrument,
            'Contract_ID': '',
            'Total_Days': inst_data['Total_Days'].sum(),
            'Loss_Days': inst_data['Loss_Days'].sum(),
            'Gain_Days': inst_data['Gain_Days'].sum(),
            'Avg_OI_Percentile_AfterLoss': inst_data['Avg_OI_Percentile_AfterLoss'].mean(),
            'Avg_NextDay_Volume_Change_AfterLoss': inst_data['Avg_NextDay_Volume_Change_AfterLoss'].mean(),
            'Avg_NextDay_OI_Normalized_AfterLoss': inst_data['Avg_NextDay_OI_Normalized_AfterLoss'].mean(),
            'Pct_OI_Increase_AfterLoss': inst_data['Pct_OI_Increase_AfterLoss'].mean(),
            'Avg_OI_Percentile_AfterGain': inst_data['Avg_OI_Percentile_AfterGain'].mean(),
            'Avg_NextDay_Volume_Change_AfterGain': inst_data['Avg_NextDay_Volume_Change_AfterGain'].mean(),
            'Avg_NextDay_OI_Normalized_AfterGain': inst_data['Avg_NextDay_OI_Normalized_AfterGain'].mean(),
            'Pct_OI_Increase_AfterGain': inst_data['Pct_OI_Increase_AfterGain'].mean(),
        }
        instrument_averages.append(avg_row)
    
    # Calculate overall average
    overall_avg = {
        'Instrument': 'OVERALL AVERAGE',
        'Base_Instrument': 'ALL',
        'Contract_ID': '',
        'Total_Days': df_report['Total_Days'].sum(),
        'Loss_Days': df_report['Loss_Days'].sum(),
        'Gain_Days': df_report['Gain_Days'].sum(),
        'Avg_OI_Percentile_AfterLoss': df_report['Avg_OI_Percentile_AfterLoss'].mean(),
        'Avg_NextDay_Volume_Change_AfterLoss': df_report['Avg_NextDay_Volume_Change_AfterLoss'].mean(),
        'Avg_NextDay_OI_Normalized_AfterLoss': df_report['Avg_NextDay_OI_Normalized_AfterLoss'].mean(),
        'Pct_OI_Increase_AfterLoss': df_report['Pct_OI_Increase_AfterLoss'].mean(),
        'Avg_OI_Percentile_AfterGain': df_report['Avg_OI_Percentile_AfterGain'].mean(),
        'Avg_NextDay_Volume_Change_AfterGain': df_report['Avg_NextDay_Volume_Change_AfterGain'].mean(),
        'Avg_NextDay_OI_Normalized_AfterGain': df_report['Avg_NextDay_OI_Normalized_AfterGain'].mean(),
        'Pct_OI_Increase_AfterGain': df_report['Pct_OI_Increase_AfterGain'].mean(),
    }
    
    df_instrument_avg = pd.DataFrame(instrument_averages)
    df_overall_avg = pd.DataFrame([overall_avg])
    
    df_final = pd.concat([df_report, df_instrument_avg, df_overall_avg], ignore_index=True)
    
    # Round numeric columns for readability
    numeric_cols = df_final.select_dtypes(include=[np.number]).columns
    df_final[numeric_cols] = df_final[numeric_cols].round(4)

    return df_final


# ==================================================
# GENERATE YEAR-WISE SUMMARY
# ==================================================
def generate_yearwise_summary(df_combined):
    """
    Generates year-wise aggregated metrics for each BASE instrument.
    Combines ALL quarterly contracts for an instrument in a given year into a SINGLE row.
    """
    df_work = df_combined.copy()
    df_work['Year'] = df_work['Date'].dt.year
    
    # Calculate percentiles for OI changes
    if "Next_Day_OI_Normalized_Change" in df_work.columns:
        df_work["Next_Day_OI_Pctl"] = (
            df_work["Next_Day_OI_Normalized_Change"].rank(pct=True) * 100
        )
    
    yearwise_data = []
    
    for (base_instrument, year), group in df_work.groupby(['Base_Instrument', 'Year']):
        loss_days = group[group['Is_Loss']]
        gain_days = group[group['Is_Gain']]
        
        # Create period description
        min_date = group['Date'].min()
        max_date = group['Date'].max()
        period_str = f"{min_date.strftime('%d %b %Y')} to {max_date.strftime('%d %b %Y')}"
        
        row = {
            'Instrument': base_instrument,
            'Year': year,
            'Period': period_str,
            'Total_Days': len(group),
            'Loss_Days': len(loss_days),
            'Gain_Days': len(gain_days),
        }
        
        # After-loss metrics
        if len(loss_days) > 0:
            row['Avg_OI_Percentile_AfterLoss'] = loss_days['Next_Day_OI_Pctl'].mean() if 'Next_Day_OI_Pctl' in loss_days.columns else np.nan
            row['Avg_NextDay_Volume_Change_AfterLoss'] = loss_days['Next_Day_Volume_Pct_Change'].mean()
            row['Avg_NextDay_OI_Normalized_AfterLoss'] = loss_days['Next_Day_OI_Normalized_Change'].mean() if 'Next_Day_OI_Normalized_Change' in loss_days.columns else np.nan
            row['Pct_OI_Increase_AfterLoss'] = (loss_days['Next_Day_OI_Change'] > 0).mean() * 100 if 'Next_Day_OI_Change' in loss_days.columns else np.nan
        else:
            row['Avg_OI_Percentile_AfterLoss'] = np.nan
            row['Avg_NextDay_Volume_Change_AfterLoss'] = np.nan
            row['Avg_NextDay_OI_Normalized_AfterLoss'] = np.nan
            row['Pct_OI_Increase_AfterLoss'] = np.nan
        
        # After-gain metrics
        if len(gain_days) > 0:
            row['Avg_OI_Percentile_AfterGain'] = gain_days['Next_Day_OI_Pctl'].mean() if 'Next_Day_OI_Pctl' in gain_days.columns else np.nan
            row['Avg_NextDay_Volume_Change_AfterGain'] = gain_days['Next_Day_Volume_Pct_Change'].mean()
            row['Avg_NextDay_OI_Normalized_AfterGain'] = gain_days['Next_Day_OI_Normalized_Change'].mean() if 'Next_Day_OI_Normalized_Change' in gain_days.columns else np.nan
            row['Pct_OI_Increase_AfterGain'] = (gain_days['Next_Day_OI_Change'] > 0).mean() * 100 if 'Next_Day_OI_Change' in gain_days.columns else np.nan
        else:
            row['Avg_OI_Percentile_AfterGain'] = np.nan
            row['Avg_NextDay_Volume_Change_AfterGain'] = np.nan
            row['Avg_NextDay_OI_Normalized_AfterGain'] = np.nan
            row['Pct_OI_Increase_AfterGain'] = np.nan
        
        yearwise_data.append(row)
    
    df_yearwise = pd.DataFrame(yearwise_data)
    df_yearwise = df_yearwise.sort_values(['Instrument', 'Year'])
    
    # Round numeric columns
    numeric_cols = df_yearwise.select_dtypes(include=[np.number]).columns
    df_yearwise[numeric_cols] = df_yearwise[numeric_cols].round(4)
    
    return df_yearwise


# ==================================================
# MAIN EXECUTION FUNCTION
# ==================================================
def run_full_analysis(output_filename=OUTPUT_FILE):
    """
    Complete pipeline: analyze all files and export to Excel with formatting.
    """
    print("\n" + "=" * 60)
    print("FUTURES MARKET MICROSTRUCTURE ANALYSIS")
    print("=" * 60 + "\n")

    # Step 1: Analyze all files in data_folder
    df_combined = analyze_data_folder()
    print(f"\nTotal rows analyzed: {len(df_combined)}")
    print(f"Unique base instruments: {df_combined['Base_Instrument'].nunique()}")
    print(f"Unique contracts: {len(df_combined.groupby(['Instrument', 'Contract_ID']))}")

    # Step 2: Generate compiled report
    print("\nGenerating compiled report...")
    df_report = generate_compiled_report(df_combined)
    
    # Step 3: Generate year-wise summary
    print("Generating year-wise summary...")
    df_yearwise = generate_yearwise_summary(df_combined)

    # Step 4: Export to Excel using openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    
    # ===== SHEET 1: Compiled Analysis =====
    ws1 = wb.active
    ws1.title = 'Compiled_Analysis'
    
    # Write headers
    headers = list(df_report.columns)
    ws1.append(headers)
    
    # Format header row
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows with conditional formatting
    instrument_avg_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    overall_avg_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    number_align = Alignment(horizontal='right', vertical='center')
    
    for idx, row in enumerate(dataframe_to_rows(df_report, index=False, header=False), start=2):
        ws1.append(row)
        
        instrument_name = str(row[0]) if row[0] is not None else ""
        
        # Apply formatting based on row type
        if "OVERALL AVERAGE" in instrument_name:
            for cell in ws1[idx]:
                cell.fill = overall_avg_fill
                cell.font = bold_font
                if cell.column > 2:
                    cell.alignment = number_align
                else:
                    cell.alignment = center_align
                    
        elif "- AVERAGE" in instrument_name:
            for cell in ws1[idx]:
                cell.fill = instrument_avg_fill
                cell.font = bold_font
                if cell.column > 2:
                    cell.alignment = number_align
                else:
                    cell.alignment = center_align
        else:
            for cell in ws1[idx]:
                cell.font = normal_font
                if cell.column > 2:
                    cell.alignment = number_align
                else:
                    cell.alignment = center_align
    
    # Set column widths
    column_widths = {
        'A': 25, 'B': 15, 'C': 12, 'D': 12, 'E': 12,
        'F': 30, 'G': 30, 'H': 25, 'I': 25,
        'J': 30, 'K': 30, 'L': 25, 'M': 25,
    }
    
    for col, width in column_widths.items():
        ws1.column_dimensions[col].width = width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Freeze top row
    ws1.freeze_panes = 'A2'
    
    # ===== SHEET 2: Year-Wise Summary =====
    ws2 = wb.create_sheet('YearWise_Summary')
    
    yearwise_headers = list(df_yearwise.columns)
    ws2.append(yearwise_headers)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for idx, row in enumerate(dataframe_to_rows(df_yearwise, index=False, header=False), start=2):
        ws2.append(row)
        
        for cell in ws2[idx]:
            cell.font = normal_font
            if cell.column > 3:
                cell.alignment = number_align
            else:
                cell.alignment = center_align
            cell.border = thin_border
    
    # Set column widths for Sheet 2
    yearwise_column_widths = {
        'A': 25, 'B': 10, 'C': 40, 'D': 12, 'E': 12, 'F': 12,
        'G': 28, 'H': 30, 'I': 30, 'J': 25,
        'K': 28, 'L': 30, 'M': 30, 'N': 25,
    }
    
    for col, width in yearwise_column_widths.items():
        ws2.column_dimensions[col].width = width
    
    ws2.freeze_panes = 'A2'
    
    # Add alternating row colors
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    current_instrument = None
    use_light_fill = False
    
    for idx in range(2, ws2.max_row + 1):
        instrument_value = ws2.cell(row=idx, column=1).value
        
        if instrument_value != current_instrument:
            current_instrument = instrument_value
            use_light_fill = not use_light_fill
        
        if use_light_fill:
            for cell in ws2[idx]:
                if cell.fill.start_color.rgb != header_fill.start_color.rgb:
                    cell.fill = light_fill
    
    # Save workbook
    wb.save(output_filename)
    full_path = os.path.abspath(output_filename)
        
    print(f"\n✓ Analysis complete!")
    print(f"✓ Report saved to: {full_path}")
    print(f"✓ Sheet 1 'Compiled_Analysis': {len(df_report)} rows")
    print(f"  - Contract-level data")
    print(f"  - Instrument-wise averages (highlighted in yellow)")
    print(f"  - Overall average (highlighted in green)")
    print(f"✓ Sheet 2 'YearWise_Summary': {len(df_yearwise)} rows")
    print(f"  - Year-wise averages for each instrument")
    print("=" * 60 + "\n")

    return df_combined, df_report, df_yearwise


# ==================================================
# RUN THE ANALYSIS
# ==================================================
if __name__ == "__main__":
    df_combined, df_report, df_yearwise = run_full_analysis()